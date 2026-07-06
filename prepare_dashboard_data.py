#!/usr/bin/env python3
"""
Prepare dashboard_data.json for the Fieldglass Onboarding Status Dashboard.

Source: "INFA contractors IQN vs FG (5).xlsx" (~/Downloads).
Status per contractor is derived from the row's Excel fill color, per the
legend provided by the business owner:

  Green  (theme 9)      -> Moved to Fieldglass
  Blue   (theme 7)       -> Transition initiated to Fieldglass
  Yellow (RGB FFFFFF00)  -> Needs clarity from manager
  Orange (theme 5)       -> Remaining in INFA system
"""
import json
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SRC = Path.home() / "Downloads" / "INFA contractors IQN vs FG (5).xlsx"
OUTPUT_PATH = Path(__file__).parent / "dashboard_data.json"

STATUS_ORDER = [
    "Moved to Fieldglass",
    "Transition initiated to Fieldglass",
    "Needs clarity from manager",
    "Remaining in INFA system",
]

STATUS_COLORS = {
    "Moved to Fieldglass": "#4CAF50",
    "Transition initiated to Fieldglass": "#2196F3",
    "Needs clarity from manager": "#FFC107",
    "Remaining in INFA system": "#FF9800",
    "Unknown": "#9E9E9E",
}


def classify(cell):
    fg = cell.fill.fgColor
    if fg.type == "rgb" and fg.rgb == "FFFFFF00":
        return "Needs clarity from manager"
    if fg.type == "theme":
        if fg.theme == 9:
            return "Moved to Fieldglass"
        if fg.theme == 7:
            return "Transition initiated to Fieldglass"
        if fg.theme == 5:
            return "Remaining in INFA system"
    return "Unknown"


def fmt_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def counter_table(grouped):
    table = []
    for key, counts in grouped.items():
        row = {"name": key, "total": sum(counts.values())}
        for s in STATUS_ORDER:
            row[s] = counts.get(s, 0)
        row["Unknown"] = counts.get("Unknown", 0)
        table.append(row)
    return table


def main():
    print(f"Reading workbook from: {SRC}")
    if not SRC.exists():
        print(f"ERROR: workbook not found at {SRC}")
        return

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = rows[1:]
    idx = {name: header.index(name) for name in header}

    contractors = []
    status_counts = Counter()
    by_manager = defaultdict(Counter)
    by_supplier = defaultdict(Counter)
    by_dept = defaultdict(Counter)
    data_issues = []

    for i, r in enumerate(data_rows):
        excel_row = i + 2
        cell = ws.cell(row=excel_row, column=1)
        status = classify(cell)
        status_counts[status] += 1

        manager = r[idx["Hiring Manager"]] or "(none)"
        supplier = r[idx["Supplier Organization"]] or "(none)"
        dept = r[idx["Department"]] or "(none)"
        by_manager[manager][status] += 1
        by_supplier[supplier][status] += 1
        by_dept[dept][status] += 1

        id_version = r[idx["ID - Version"]]
        if id_version == "#VALUE!":
            data_issues.append({"row": excel_row, "resource": r[idx["Resource"]]})

        contractors.append({
            "id_version": str(id_version) if id_version is not None else "",
            "resource": r[idx["Resource"]] or "",
            "status": status,
            "current_phase": r[idx["Current Phase"]] or "",
            "start_date": fmt_date(r[idx["Start Date"]]),
            "end_date": fmt_date(r[idx["End Date"]]),
            "hiring_manager": manager,
            "position_title": r[idx["Position Title"]] or "",
            "location": r[idx["Location"]] or "",
            "purchase_order": str(r[idx["Purchase Order"]]) if r[idx["Purchase Order"]] is not None else "",
            "supplier_organization": supplier,
            "department": dept,
            "email": r[idx["Resource Email ID"]] or "",
            "emp_id": str(r[idx["Emp ID"]]) if r[idx["Emp ID"]] is not None else "",
        })

    total = len(contractors)
    assert sum(status_counts.values()) == total
    print(f"Loaded {total} contractor records")
    for s in STATUS_ORDER:
        print(f"  {s}: {status_counts.get(s, 0)}")
    if status_counts.get("Unknown"):
        print(f"  Unknown: {status_counts['Unknown']}")

    manager_table = sorted(
        counter_table(by_manager),
        key=lambda r: (-r["Needs clarity from manager"], -r["total"]),
    )
    supplier_table = sorted(counter_table(by_supplier), key=lambda r: -r["total"])
    dept_table = sorted(counter_table(by_dept), key=lambda r: -r["total"])

    output_data = {
        "generated_from": "INFA contractors IQN vs FG (5).xlsx",
        "total_contractors": total,
        "status_order": STATUS_ORDER,
        "status_colors": STATUS_COLORS,
        "status_counts": {s: status_counts.get(s, 0) for s in STATUS_ORDER},
        "unknown_count": status_counts.get("Unknown", 0),
        "data_issues": data_issues,
        "by_manager": manager_table,
        "by_supplier": supplier_table,
        "by_department": dept_table,
        "contractors": contractors,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    print(f"\nDashboard data written to: {OUTPUT_PATH}")

    print("\n=== Top 10 managers by 'Needs clarity from manager' count ===")
    for m in manager_table[:10]:
        if m["Needs clarity from manager"] > 0:
            print(f"  {m['name'][:40]:40s} {m['Needs clarity from manager']:3d} needs clarity (of {m['total']:3d} total)")


if __name__ == "__main__":
    main()
